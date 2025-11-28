"""FastAPI dashboard for exploring Lowe's clearance items."""

from __future__ import annotations

import json
import os
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
import time
import re
from typing import Any, Iterable, Literal
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

from fastapi import Depends, FastAPI, HTTPException, Query, Request
from fastapi.responses import JSONResponse, PlainTextResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.logging_config import get_logger
from app.middleware.simple_session import SimpleSessionMiddleware
from app.lowes_stores_wa_or import LOWES_STORES_WA_OR
from app.storage import repo
from app.storage.db import get_engine, init_db, make_session


LOGGER = get_logger(__name__)
BASE_PATH = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_PATH / "templates"
STATIC_DIR = BASE_PATH / "static"
DATABASE_FILE = Path(__file__).resolve().parent.parent / "orwa_lowes.sqlite"

DB_BUSY_TIMEOUT = float(os.getenv("DB_BUSY_TIMEOUT", "30"))
engine = get_engine(str(DATABASE_FILE), busy_timeout=DB_BUSY_TIMEOUT)
init_db(engine)
session_factory = make_session(engine)

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
app = FastAPI(title="CheapSkater Clearance Dashboard")
SESSION_SECRET = os.getenv("CHEAPSKATER_SESSION_SECRET", "cheapskater-session-secret")
SESSION_MAX_AGE = int(os.getenv("CHEAPSKATER_SESSION_MAX_AGE", str(60 * 60 * 24 * 30)))
app.add_middleware(SimpleSessionMiddleware, secret_key=SESSION_SECRET, max_age=SESSION_MAX_AGE)
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

_STATS_CACHE: dict[str, Any] | None = None
_STATS_CACHE_TS: float | None = None

@app.exception_handler(404)
async def not_found_handler(request: Request, exc: HTTPException) -> PlainTextResponse:
    if STATIC_DIR.exists():
        detail = exc.detail if isinstance(exc.detail, str) else "Not Found"
        return PlainTextResponse(detail, status_code=404)
    return PlainTextResponse(
        "CheapSkater Dashboard UI unavailable.\n"
        "Use: /api/clearance, /api/stats, /healthz",
        status_code=404,
    )

Scope = Literal["all", "new"]
STATE_OPTIONS = ["ALL", "WA", "OR"]
DEFAULT_CATEGORY_OPTIONS = [
    "Roofing",
    "Drywall",
    "Insulation",
    "Lumber",
    "Flooring",
    "Plumbing",
    "Electrical",
    "Tools",
    "Paint",
    "Hardware",
]
INITIAL_GROUP_BATCH = 30
_STORE_HOURS_SUFFIX = re.compile(r"\s+\d{1,2}\s*(?:AM|PM)$", re.I)
_STOCK_COUNT_PATTERN = re.compile(r"(?:only|about)?\s*(\d+)\s*(?:left|available|in\s*stock|qty|quantity)", re.I)
_SAVED_DEALS_SESSION_KEY = "saved_deals"
_STORE_STATUS_TOKEN_PATTERN = re.compile(r"\s+(?:closed|open)\b", re.I)
_STORE_HOURS_RANGE_PATTERN = re.compile(
    r"\s+\d{1,2}\s*(?:AM|PM)\s*(?:[-–—]\s*\d{1,2}\s*(?:AM|PM))?",
    re.I,
)
_STORE_HOURS_VERB_PATTERN = re.compile(r"\s+(?:opens|closes)\s+\d{1,2}\s*(?:AM|PM)", re.I)


class SaveDealPayload(BaseModel):
    store_number: str = Field(..., description="Lowe's store number.")
    sku: str = Field(..., description="Product SKU.")
    quantity: int | None = Field(
        default=1, ge=1, description="Quantity to add to the saved cart."
    )


def _normalize_store_number(value: str | None) -> str:
    """Return a zero-padded store number for lookups."""

    text = (value or "").strip()
    if text.isdigit() and len(text) < 4:
        return text.zfill(4)
    return text


def _strip_store_status_text(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip()
    if not text:
        return None
    text = _STORE_STATUS_TOKEN_PATTERN.sub("", text)
    text = _STORE_HOURS_VERB_PATTERN.sub("", text)
    text = _STORE_HOURS_RANGE_PATTERN.sub("", text)
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip() or None


def normalize_store_label(store_number: str | None, store_name_raw: str | None) -> str:
    """Resolve a user-facing store label, preferring the canonical address."""

    normalized_number = _normalize_store_number(store_number)
    details = LOWES_STORES_WA_OR.get(normalized_number)
    if details:
        return details["address"]
    cleaned = _strip_store_status_text(store_name_raw)
    if cleaned:
        return cleaned
    if normalized_number:
        return f"Lowe's #{normalized_number}"
    return "Lowe's Store"


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _coerce_datetime(value: object | None) -> datetime | None:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)
    return None


def _relative_days(value: object | None) -> int | None:
    dt_value = _coerce_datetime(value)
    if not dt_value:
        return None
    delta = _now_utc() - dt_value
    return max(int(delta.total_seconds() // 86400), 0)


def _time_delta_for(value: str | None) -> timedelta | None:
    for key, _, delta in TIME_FILTER_OPTIONS:
        if key == value:
            return delta
    return None


def _sanitize_percent(value: float | None) -> float | None:
    if value is None:
        return None
    return max(0.0, min(100.0, float(value)))


def _sanitize_stock(value: float | None) -> int | None:
    if value is None:
        return None
    return max(0, int(value))


def _as_float(value: float | str | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = value.strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _as_int(value: int | str | None) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = value.strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _estimate_stock_units(availability: str | None) -> int | None:
    if not availability:
        return None
    normalized = availability.strip().lower()
    if not normalized:
        return None
    match = _STOCK_COUNT_PATTERN.search(normalized)
    if match:
        try:
            return max(int(match.group(1)), 0)
        except (TypeError, ValueError):
            return None
    if "out of stock" in normalized:
        return 0
    if "limited" in normalized:
        return 1
    if "in stock" in normalized:
        return 5
    if "only" in normalized:
        return 1
    return None


def _format_stock_status(stock_estimate: int | None, availability: str | None) -> str | None:
    if stock_estimate is None:
        return None
    if stock_estimate == 0:
        return "0 in stock"
    units = "unit" if stock_estimate == 1 else "units"
    return f"{stock_estimate} {units} available"


def _prepare_listing(listing: dict[str, Any]) -> dict[str, Any]:
    enriched = dict(listing)
    enriched["store_label"] = _format_store_label(enriched)
    enriched["store_tooltip"] = _format_store_tooltip(enriched)
    enriched["store_product_url"] = _store_specific_url(
        enriched.get("product_url"), enriched.get("store_id")
    )
    added_ts = enriched.get("first_seen") or enriched.get("price_started_at")
    enriched["days_since_added"] = _relative_days(added_ts)
    stock_estimate = _estimate_stock_units(enriched.get("availability"))
    enriched["stock_estimate"] = stock_estimate
    enriched["stock_label"] = _format_stock_status(stock_estimate, enriched.get("availability"))
    return enriched


def _prepare_listings(listings: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return [_prepare_listing(listing) for listing in listings]


def _apply_filters(
    listings: Iterable[dict[str, Any]],
    *,
    filters: dict[str, Any],
) -> list[dict[str, Any]]:
    cutoff = filters.get("time_cutoff")
    discount_min = filters.get("discount_min")
    stock_min = filters.get("stock_min")
    stock_max = filters.get("stock_max")

    result: list[dict[str, Any]] = []
    for listing in listings:
        added_at = listing.get("first_seen") or listing.get("price_started_at")
        added_dt = _coerce_datetime(added_at)
        if cutoff and (added_dt is None or added_dt < cutoff):
            continue
        pct_off = listing.get("pct_off")
        if discount_min is not None:
            if pct_off is None or pct_off < discount_min:
                continue
        stock_estimate = listing.get("stock_estimate")
        if stock_min is not None:
            if stock_estimate is None or stock_estimate < stock_min:
                continue
        if stock_max is not None:
            if stock_estimate is None or stock_estimate > stock_max:
                continue
        result.append(listing)
    return result


def _normalize_filters(
    *,
    time_window: str | None,
    discount_filter: str | None,
    discount_min: str | float | None,
    stock_filter: str | None,
    stock_min: str | int | None,
    stock_max: str | int | None,
) -> dict[str, Any]:
    normalized_time = time_window if any(opt[0] == time_window for opt in TIME_FILTER_OPTIONS) else "all"
    delta = _time_delta_for(normalized_time)
    cutoff = None
    if delta:
        cutoff = _now_utc() - delta

    discount_choice = discount_filter or "all"
    preset_discounts = {
        value: min_value
        for value, _, min_value, _ in DISCOUNT_FILTER_OPTIONS
        if min_value is not None and value not in {"all", "custom"}
    }
    min_pct = None
    if discount_choice in preset_discounts:
        min_pct = preset_discounts[discount_choice]
    elif discount_choice == "custom":
        parsed_min = _sanitize_percent(_as_float(discount_min))
        min_pct = parsed_min
    else:
        discount_choice = "all"

    stock_choice = stock_filter or "all"
    preset_stock = {
        value: min_value
        for value, _, min_value, _ in STOCK_FILTER_OPTIONS
        if min_value is not None and value not in {"all", "custom"}
    }
    stock_floor = None
    stock_ceiling = None
    if stock_choice in preset_stock:
        stock_floor = preset_stock[stock_choice]
    elif stock_choice == "custom":
        parsed_min = _as_int(stock_min)
        parsed_max = _as_int(stock_max)
        stock_floor = _sanitize_stock(parsed_min)
        stock_ceiling = _sanitize_stock(parsed_max)
        if stock_floor is not None and stock_ceiling is not None and stock_ceiling < stock_floor:
            stock_floor, stock_ceiling = stock_ceiling, stock_floor
    else:
        stock_choice = "all"

    return {
        "time_window": normalized_time,
        "time_cutoff": cutoff,
        "discount_choice": discount_choice,
        "discount_min_pct": min_pct,
        "discount_min": (min_pct / 100.0) if min_pct is not None else None,
        "stock_choice": stock_choice,
        "stock_min": stock_floor,
        "stock_max": stock_ceiling,
    }


def _spread(min_value: float | None, max_value: float | None) -> float | None:
    if min_value is None or max_value is None:
        return None
    delta = max_value - min_value
    return delta if delta > 0 else 0.0


TIME_FILTER_OPTIONS = [
    ("all", "All Dates", None),
    ("1h", "Last 1 hour", timedelta(hours=1)),
    ("12h", "Last 12 hours", timedelta(hours=12)),
    ("24h", "Last 24 hours", timedelta(days=1)),
    ("3d", "Last 3 days", timedelta(days=3)),
    ("1w", "Last 1 week", timedelta(weeks=1)),
    ("1m", "Last 1 month", timedelta(days=30)),
]
DISCOUNT_FILTER_OPTIONS = [
    ("all", "All Discounts", None, None),
    ("60", "60%+", 60.0, None),
    ("75", "75%+", 75.0, None),
    ("90", "90%+", 90.0, None),
    ("custom", "Custom Range…", None, None),
]
STOCK_FILTER_OPTIONS = [
    ("all", "Any Stock", None, None),
    ("1", "1+ Stock", 1, None),
    ("2", "2+ Stock", 2, None),
    ("3", "3+ Stock", 3, None),
    ("4", "4+ Stock", 4, None),
    ("5", "5+ Stock", 5, None),
    ("custom", "Custom Range…", None, None),
]


def _store_specific_url(url: str | None, store_id: str | None) -> str | None:
    if not url:
        return None
    if not store_id:
        return url

    parsed = urlparse(url)
    params = parse_qsl(parsed.query, keep_blank_values=True)
    has_store = any(key.lower() == "storenumber" for key, _ in params)
    if not has_store:
        params.append(("storeNumber", store_id))
    new_query = urlencode(params, doseq=True)
    return urlunparse(parsed._replace(query=new_query))


def _clean_store_name(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip()
    if not text:
        return None
    return _STORE_HOURS_SUFFIX.sub("", text)


def _format_store_label(listing: dict[str, Any]) -> str:
    name = _clean_store_name(listing.get("store_name"))
    city = (listing.get("store_city") or "").strip()
    state = (listing.get("store_state") or "").strip()
    store_id = (listing.get("store_id") or "").strip()

    label = name or ""
    if not label and city:
        label = f"{city} Lowe's"
    if not label and store_id:
        label = f"Lowe's #{store_id}"
    if not label:
        label = "Lowe's Store"
    if city and city.lower() not in label.lower():
        label = f"{label} – {city}"
    elif state and state not in label:
        label = f"{label} ({state})"
    return label


def _format_store_tooltip(listing: dict[str, Any]) -> str:
    city = (listing.get("store_city") or "").strip()
    state = (listing.get("store_state") or "").strip()
    zip_code = (listing.get("store_zip") or "").strip()
    store_id = (listing.get("store_id") or "").strip()
    parts = [part for part in [city, state, zip_code] if part]
    details = ", ".join(parts)
    if store_id:
        details = f"{details} • Store #{store_id}" if details else f"Store #{store_id}"
    return details or "Store location"


def get_session() -> Iterable[Session]:
    """Dependency that yields a SQLAlchemy session."""

    with session_factory() as session:
        yield session


def _cache_stats(session: Session) -> dict[str, Any]:
    global _STATS_CACHE, _STATS_CACHE_TS
    now = time.monotonic()
    if _STATS_CACHE and _STATS_CACHE_TS and (now - _STATS_CACHE_TS) < 60:
        return _STATS_CACHE
    total_items = repo.count_observations(session)
    quarantine_total = repo.count_quarantine(session)
    last_scrape = repo.get_latest_timestamp(session)
    payload = {
        "total_items": total_items,
        "quarantine_count": quarantine_total,
        "last_scrape": last_scrape.isoformat() if last_scrape else None,
    }
    _STATS_CACHE, _STATS_CACHE_TS = payload, now
    return payload


def _serialize_observation(listing: dict[str, Any]) -> dict[str, Any]:
    """Convert a store listing mapping into JSON-serialisable data."""

    return _serialize_listing(listing)


def _format_timestamp(value: Any, *, show_time: bool = True) -> str | None:
    if not value:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        try:
            value = datetime.fromisoformat(text)
        except ValueError:
            return text
    if not isinstance(value, datetime):
        return str(value)
    fmt = "%b %d %I:%M %p" if show_time else "%b %d"
    return value.strftime(fmt)


def _format_currency(value: float | None) -> str | None:
    if value is None:
        return None
    return f"${value:,.2f}"


def _normalize_state(value: str | None) -> str | None:
    if not value:
        return None
    upper = value.upper()
    return upper if upper in {"WA", "OR"} else None


def _state_from_zip(zip_code: str | None) -> str | None:
    if not zip_code:
        return None
    digits = "".join(ch for ch in zip_code if ch.isdigit())
    if len(digits) < 3:
        return None
    prefix = int(digits[:3])
    if 970 <= prefix <= 979:
        return "OR"
    if 980 <= prefix <= 994:
        return "WA"
    return None


def _serialize_listing(listing: dict[str, Any]) -> dict[str, Any]:
    def _ts(value: Any) -> str | None:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, str):
            text = value.strip()
            return text or None
        return value

    payload = {
        "history_id": listing.get("history_id"),
        "retailer": listing.get("retailer"),
        "store_id": listing.get("store_id"),
        "store_name": listing.get("store_name"),
        "store_city": listing.get("store_city"),
        "store_state": listing.get("store_state"),
        "store_zip": listing.get("store_zip"),
        "sku": listing.get("sku"),
        "title": listing.get("title"),
        "category": listing.get("category"),
        "price": listing.get("price"),
        "price_was": listing.get("price_was"),
        "pct_off": listing.get("pct_off"),
        "availability": listing.get("availability"),
        "product_url": listing.get("product_url"),
        "store_product_url": _store_specific_url(
            listing.get("product_url"), listing.get("store_id")
        ),
        "image_url": listing.get("image_url"),
        "clearance": listing.get("clearance"),
        "first_seen": _ts(listing.get("first_seen")),
        "price_started_at": _ts(listing.get("price_started_at")),
        "updated_at": _ts(listing.get("updated_at")),
        "prev_price": listing.get("prev_price"),
        "prev_price_was": listing.get("prev_price_was"),
        "prev_pct_off": listing.get("prev_pct_off"),
        "prev_updated_at": _ts(listing.get("prev_updated_at")),
        "prev_clearance": listing.get("prev_clearance"),
        "store_label": _format_store_label(listing),
        "store_tooltip": _format_store_tooltip(listing),
        "store_product_url": listing.get("store_product_url")
        or _store_specific_url(listing.get("product_url"), listing.get("store_id")),
        "stock_estimate": listing.get("stock_estimate"),
        "stock_label": listing.get("stock_label"),
        "days_since_added": _relative_days(listing.get("first_seen") or listing.get("price_started_at")),
    }
    return payload


def _group_listings(listings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for listing in listings:
        sku = listing.get("sku") or listing.get("history_id")
        if not sku:
            continue
        bucket = grouped.setdefault(
            sku,
            {
                "sku": sku,
                "title": listing.get("title"),
                "category": listing.get("category"),
                "image_url": listing.get("image_url"),
                "stores": [],
                "added_at": listing.get("first_seen") or listing.get("price_started_at"),
                "last_seen": listing.get("updated_at") or listing.get("price_started_at"),
                "days_since_added": listing.get("days_since_added"),
            },
        )
        if not bucket["title"] and listing.get("title"):
            bucket["title"] = listing.get("title")
        if not bucket["category"] and listing.get("category"):
            bucket["category"] = listing.get("category")
        if not bucket["image_url"] and listing.get("image_url"):
            bucket["image_url"] = listing.get("image_url")
        listing_added = listing.get("first_seen") or listing.get("price_started_at")
        if listing_added and (
            not bucket["added_at"]
            or (
                isinstance(listing_added, datetime)
                and isinstance(bucket["added_at"], datetime)
                and listing_added < bucket["added_at"]
            )
        ):
            bucket["added_at"] = listing_added
        listing_seen = listing.get("updated_at") or listing.get("price_started_at")
        if listing_seen and (
            not bucket["last_seen"]
            or (
                isinstance(listing_seen, datetime)
                and isinstance(bucket["last_seen"], datetime)
                and listing_seen > bucket["last_seen"]
            )
        ):
            bucket["last_seen"] = listing_seen
        bucket["stores"].append(listing)

    grouped_list: list[dict[str, Any]] = []
    for bucket in grouped.values():
        stores = bucket["stores"]
        stores.sort(
            key=lambda row: (
                row.get("price") if row.get("price") is not None else float("inf"),
                row.get("pct_off") if row.get("pct_off") is not None else 0,
            )
        )
        prices = [row.get("price") for row in stores if row.get("price") is not None]
        discounts = [row.get("pct_off") for row in stores if row.get("pct_off") is not None]
        savings = []
        for row in stores:
            price = row.get("price")
            price_was = row.get("price_was")
            if price is None or price_was is None:
                continue
            savings.append(max(price_was - price, 0))
        stock_values = [
            row.get("stock_estimate") for row in stores if row.get("stock_estimate") is not None
        ]
        bucket["min_price"] = min(prices) if prices else None
        bucket["max_price"] = max(prices) if prices else None
        bucket["min_pct_off"] = min(discounts) if discounts else None
        bucket["max_pct_off"] = max(discounts) if discounts else None
        bucket["price_spread"] = _spread(bucket["min_price"], bucket["max_price"])
        bucket["discount_spread"] = _spread(bucket["min_pct_off"], bucket["max_pct_off"])
        bucket["min_stock_estimate"] = min(stock_values) if stock_values else None
        bucket["max_stock_estimate"] = max(stock_values) if stock_values else None
        bucket["min_savings"] = min(savings) if savings else None
        bucket["max_savings"] = max(savings) if savings else None
        bucket["locations"] = len(stores)
        if stores:
            best_store = stores[0]
            bucket["best_product_url"] = best_store.get("store_product_url") or best_store.get(
                "product_url"
            )
        else:
            bucket["best_product_url"] = None
        bucket["days_since_added"] = bucket.get("days_since_added") or _relative_days(
            bucket.get("added_at")
        )
        grouped_list.append(bucket)

    grouped_list.sort(
        key=lambda row: (
            row.get("min_price") if row.get("min_price") is not None else float("inf"),
            row.get("sku"),
        )
    )
    return grouped_list


def _serialize_group(group: dict[str, Any]) -> dict[str, Any]:
    price_was_values = [
        store.get("price_was")
        for store in group.get("stores", [])
        if store.get("price_was") is not None
    ]

    payload = {
        "sku": group.get("sku"),
        "title": group.get("title"),
        "category": group.get("category"),
        "image_url": group.get("image_url"),
        "min_price": group.get("min_price"),
        "max_price": group.get("max_price"),
        "min_price_was": min(price_was_values) if price_was_values else None,
        "max_price_was": max(price_was_values) if price_was_values else None,
        "min_pct_off": group.get("min_pct_off"),
        "max_pct_off": group.get("max_pct_off"),
        "price_spread": group.get("price_spread"),
        "discount_spread": group.get("discount_spread"),
        "min_stock_estimate": group.get("min_stock_estimate"),
        "max_stock_estimate": group.get("max_stock_estimate"),
        "min_savings": group.get("min_savings"),
        "max_savings": group.get("max_savings"),
        "locations": group.get("locations"),
        "added_at": group.get("added_at").isoformat() if isinstance(group.get("added_at"), datetime) else group.get("added_at"),
        "last_seen": group.get("last_seen").isoformat() if isinstance(group.get("last_seen"), datetime) else group.get("last_seen"),
        "days_since_added": group.get("days_since_added") or _relative_days(group.get("added_at")),
        "best_product_url": group.get("best_product_url"),
        "stores": [_serialize_listing(store) for store in group.get("stores", [])],
    }
    return payload


def _coerce_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _coerce_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _datetime_to_iso(value: object | None) -> str | None:
    if isinstance(value, datetime):
        target = value
        if target.tzinfo is None:
            target = target.replace(tzinfo=timezone.utc)
        else:
            target = target.astimezone(timezone.utc)
        return target.isoformat()
    return None


def _build_cheapskater_deal(listing: dict[str, Any]) -> dict[str, Any]:
    store_number_raw = (listing.get("store_id") or "").strip()
    store_number = _normalize_store_number(store_number_raw) or store_number_raw
    sku = (listing.get("sku") or "").strip()
    deal_id = f"{store_number}:{sku}" if store_number or sku else sku
    price_value = _coerce_float(listing.get("price"))
    price_was_value = listing.get("price_was")
    price_was = _coerce_float(price_was_value) if price_was_value is not None else None
    pct_off_raw = _coerce_float(listing.get("pct_off"))
    pct_off_value = None
    if pct_off_raw is not None:
        pct_off_value = pct_off_raw if pct_off_raw > 1 else pct_off_raw * 100
    stock_estimate = listing.get("stock_estimate")
    stock_value = _coerce_int(stock_estimate)
    store_name_raw = listing.get("store_name") or ""
    last_updated = listing.get("updated_at") or listing.get("price_started_at")
    last_updated_label = _format_timestamp(last_updated)
    return {
        "deal_id": deal_id,
        "store_number": store_number,
        "store_name_raw": store_name_raw,
        "store_label": normalize_store_label(store_number, store_name_raw),
        "store_city": listing.get("store_city"),
        "store_state": listing.get("store_state"),
        "store_zip": listing.get("store_zip"),
        "state": listing.get("store_state"),
        "zip": listing.get("store_zip"),
        "product_name": listing.get("title"),
        "sku": sku,
        "category": listing.get("category"),
        "price": price_value,
        "price_display": _format_currency(price_value),
        "price_was": price_was,
        "price_was_display": _format_currency(price_was) if price_was is not None else None,
        "pct_off": pct_off_value,
        "pct_off_label": f"{pct_off_value:.0f}% off" if pct_off_value is not None else None,
        "availability": listing.get("availability"),
        "stock": stock_value if stock_value is not None else 0,
        "stock_estimate": stock_estimate,
        "stock_label": listing.get("stock_label"),
        "last_updated": _datetime_to_iso(last_updated),
        "last_updated_label": last_updated_label,
        "days_since_added": listing.get("days_since_added"),
        "product_url": listing.get("product_url"),
        "image_url": listing.get("image_url"),
        "price_started_at": listing.get("price_started_at"),
    }


def _lookup_listing_for_save(
    session: Session,
    store_number: str,
    sku: str,
) -> dict[str, Any] | None:
    candidates: list[str] = []
    raw = (store_number or "").strip()
    if raw:
        candidates.append(raw)
    normalized = _normalize_store_number(raw)
    if normalized and normalized not in candidates:
        candidates.append(normalized)
    trimmed = normalized.lstrip("0") if normalized else raw.lstrip("0")
    if trimmed and trimmed not in candidates:
        candidates.append(trimmed)
    for candidate in candidates:
        if not candidate:
            continue
        listing = repo.get_listing_for_store_and_sku(
            session, store_id=candidate, sku=sku
        )
        if listing:
            return listing
    return None


def _get_saved_deals(request: Request) -> dict[str, dict[str, Any]]:
    stored = request.session.get(_SAVED_DEALS_SESSION_KEY)
    if isinstance(stored, dict):
        return dict(stored)
    request.session[_SAVED_DEALS_SESSION_KEY] = {}
    return {}


def _persist_saved_deals(request: Request, saved: dict[str, dict[str, Any]]) -> None:
    request.session[_SAVED_DEALS_SESSION_KEY] = saved


def _saved_entry_from_deal(deal: dict[str, Any], quantity: int) -> dict[str, Any]:
    return {
        "deal_id": deal.get("deal_id"),
        "store_number": deal.get("store_number"),
        "store_label": deal.get("store_label"),
        "product_name": deal.get("product_name"),
        "sku": deal.get("sku"),
        "price": deal.get("price"),
        "stock": deal.get("stock"),
        "quantity": quantity,
        "last_updated": deal.get("last_updated"),
        "product_url": deal.get("product_url"),
    }


def _cart_totals(saved: dict[str, dict[str, Any]]) -> tuple[int, float]:
    total_items = 0
    total_price = 0.0
    for entry in saved.values():
        qty = int(entry.get("quantity", 0) or 0)
        price = _coerce_float(entry.get("price"))
        if qty < 0:
            qty = 0
        total_items += qty
        total_price += price * qty
    return total_items, round(total_price, 2)


def _deal_line_total(entry: dict[str, Any]) -> float:
    qty = int(entry.get("quantity", 0) or 0)
    if qty < 0:
        qty = 0
    price = _coerce_float(entry.get("price"))
    return round(price * qty, 2)


def _serialize_saved_deal(entry: dict[str, Any]) -> dict[str, Any]:
    payload = dict(entry)
    payload["line_total"] = _deal_line_total(entry)
    return payload


def _group_saved_deals(saved: dict[str, dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in saved.values():
        store_number = entry.get("store_number") or "unknown"
        grouped[store_number].append(entry)
    return dict(grouped)


def _select_items(
    session: Session,
    *,
    scope: Scope,
    state: str | None,
    category: str | None,
) -> list[dict[str, Any]]:
    if scope == "new":
        return repo.get_new_clearance_today(session, state=state, category=category)
    return repo.get_clearance_items(session, state=state, category=category)


def _collect_categories(session: Session) -> list[str]:
    discovered = repo.list_distinct_categories(session)
    merged = sorted({*DEFAULT_CATEGORY_OPTIONS, *discovered})
    return merged


@app.get("/cheapskater")
def cheapskater_view(request: Request):
    """Render the Cheapskater deal board with the session-backed cart."""

    saved = _get_saved_deals(request)
    cart_items = list(saved.values())
    cart_total_items, cart_total_price = _cart_totals(saved)
    grouped = _group_saved_deals(saved)
    cart_groups = [
        {
            "store_number": store_number,
            "store_label": entries[0].get("store_label")
            or normalize_store_label(store_number, None),
            "deals": entries,
        }
        for store_number, entries in grouped.items()
    ]
    return templates.TemplateResponse(
        "cheapskater.html",
        {
            "request": request,
            "cart_items": cart_items,
            "cart_groups": cart_groups,
            "cart_total_items": cart_total_items,
            "cart_total_price": cart_total_price,
            "format_currency": _format_currency,
            "active_scope": "cheapskater",
            "state": "ALL",
            "has_saved": bool(cart_items),
        },
    )


@app.post("/cheapskater/save-deal")
def save_deal(
    payload: SaveDealPayload,
    request: Request,
    session: Session = Depends(get_session),
) -> JSONResponse:
    """Persist or update a saved deal entry in the session cart."""

    store_number = (payload.store_number or "").strip()
    sku = (payload.sku or "").strip()
    if not store_number or not sku:
        raise HTTPException(status_code=400, detail="store_number and sku are required.")
    listing = _lookup_listing_for_save(session, store_number, sku)
    if listing is None:
        raise HTTPException(status_code=404, detail="Deal not found.")
    enriched = _prepare_listing(listing)
    deal = _build_cheapskater_deal(enriched)
    saved = _get_saved_deals(request)
    deal_id = deal.get("deal_id") or f"{store_number}:{sku}"
    quantity = int(payload.quantity or 1)
    entry = saved.get(deal_id)
    if entry:
        entry["quantity"] = int(entry.get("quantity", 0) or 0) + quantity
        entry["price"] = deal.get("price")
        entry["stock"] = deal.get("stock")
        entry["store_label"] = deal.get("store_label")
        entry["product_name"] = deal.get("product_name")
        entry["last_updated"] = deal.get("last_updated")
        entry["product_url"] = deal.get("product_url")
    else:
        entry = _saved_entry_from_deal(deal, quantity)
    saved[deal_id] = entry
    _persist_saved_deals(request, saved)
    cart_total_items, cart_total_price = _cart_totals(saved)
    return JSONResponse(
        content={
            "ok": True,
            "deal_id": deal_id,
            "cart_count": cart_total_items,
            "cart_total": cart_total_price,
            "deal": _serialize_saved_deal(entry),
        }
    )


@app.post("/cheapskater/cart/{deal_id}/increment")
def increment_saved_deal(deal_id: str, request: Request) -> JSONResponse:
    """Increase the planned quantity for a saved deal."""

    saved = _get_saved_deals(request)
    entry = saved.get(deal_id)
    if entry is None:
        raise HTTPException(status_code=404, detail="Deal not stored in cart.")
    entry["quantity"] = int(entry.get("quantity", 0) or 0) + 1
    saved[deal_id] = entry
    _persist_saved_deals(request, saved)
    cart_total_items, cart_total_price = _cart_totals(saved)
    return JSONResponse(
        content={
            "ok": True,
            "deal_id": deal_id,
            "quantity": entry["quantity"],
            "cart_count": cart_total_items,
            "cart_total": cart_total_price,
            "line_total": _deal_line_total(entry),
            "store_number": entry.get("store_number"),
            "store_label": entry.get("store_label"),
        }
    )


@app.post("/cheapskater/cart/{deal_id}/decrement")
def decrement_saved_deal(deal_id: str, request: Request) -> JSONResponse:
    """Decrease the planned quantity for a saved deal (removing when zero)."""

    saved = _get_saved_deals(request)
    entry = saved.get(deal_id)
    if entry is None:
        raise HTTPException(status_code=404, detail="Deal not stored in cart.")
    current_quantity = int(entry.get("quantity", 0) or 0) - 1
    deleted = current_quantity <= 0
    store_number = entry.get("store_number")
    store_label = entry.get("store_label")
    if deleted:
        saved.pop(deal_id, None)
    else:
        entry["quantity"] = current_quantity
        saved[deal_id] = entry
    _persist_saved_deals(request, saved)
    cart_total_items, cart_total_price = _cart_totals(saved)
    return JSONResponse(
        content={
            "ok": True,
            "deal_id": deal_id,
            "quantity": current_quantity if not deleted else 0,
            "cart_count": cart_total_items,
            "cart_total": cart_total_price,
            "line_total": _deal_line_total(entry) if not deleted else 0.0,
            "store_number": store_number,
            "store_label": store_label,
            "deleted": deleted,
        }
    )


def _render_dashboard(
    request: Request,
    *,
    scope: Scope,
    state: str | None,
    category: str | None,
    filters: dict[str, Any],
    session: Session,
):
    LOGGER.debug(
        "Rendering dashboard", extra={"scope": scope, "state": state, "category": category}
    )
    raw_items = _select_items(session, scope=scope, state=state, category=category)
    prepared_items = _prepare_listings(raw_items)
    items = _apply_filters(prepared_items, filters=filters)
    grouped = _group_listings(items)
    serialized_groups = [_serialize_group(group) for group in grouped]
    initial_groups = serialized_groups[:INITIAL_GROUP_BATCH]
    LOGGER.info(
        "Dashboard payload | groups=%s initial=%s",
        len(serialized_groups),
        len(initial_groups),
    )
    categories = _collect_categories(session)
    last_updated = repo.get_latest_timestamp(session)

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "items": items,
            "groups": serialized_groups,
            "initial_groups": initial_groups,
            "group_json": json.dumps(serialized_groups, ensure_ascii=False),
            "initial_batch_size": INITIAL_GROUP_BATCH,
            "active_scope": scope,
            "state": state or "ALL",
            "category": category,
            "categories": categories,
            "state_options": STATE_OPTIONS,
            "last_updated": last_updated,
            "state_from_zip": _state_from_zip,
            "store_url_builder": _store_specific_url,
            "store_label_builder": _format_store_label,
            "store_tooltip_builder": _format_store_tooltip,
            "format_timestamp": _format_timestamp,
            "format_currency": _format_currency,
            "filters": filters,
            "time_filter_options": TIME_FILTER_OPTIONS,
            "discount_filter_options": DISCOUNT_FILTER_OPTIONS,
            "stock_filter_options": STOCK_FILTER_OPTIONS,
        },
    )


@app.get("/healthz")
def healthcheck() -> dict[str, str]:
    """Return application health information."""

    return {"status": "ok"}


@app.get("/")
def list_clearance(
    request: Request,
    state: str | None = Query(None, description="State filter (WA or OR)."),
    category: str | None = Query(None, description="Optional category filter."),
    time_window: str = Query("all", description="Time filter window key."),
    discount_filter: str | None = Query(
        None, description="Discount preset (percentage or custom)."
    ),
    discount_min: str | None = Query(
        None, description="Custom minimum discount percentage."
    ),
    stock_filter: str | None = Query(
        None, description="Stock preset (quantity or custom)."
    ),
    stock_min: str | None = Query(None, description="Custom minimum stock value."),
    stock_max: str | None = Query(None, description="Custom maximum stock value."),
    session: Session = Depends(get_session),
):
    """Render the full clearance dashboard."""

    normalized_state = _normalize_state(state)
    normalized_category = category or None
    filters = _normalize_filters(
        time_window=time_window,
        discount_filter=discount_filter,
        discount_min=discount_min,
        stock_filter=stock_filter,
        stock_min=stock_min,
        stock_max=stock_max,
    )

    return _render_dashboard(
        request,
        scope="all",
        state=normalized_state,
        category=normalized_category,
        filters=filters,
        session=session,
    )


@app.get("/new-today")
def list_new_clearance_today(
    request: Request,
    state: str | None = Query(None, description="State filter (WA or OR)."),
    category: str | None = Query(None, description="Optional category filter."),
    time_window: str = Query("all", description="Time filter window key."),
    discount_filter: str | None = Query(
        None, description="Discount preset (percentage or custom)."
    ),
    discount_min: str | None = Query(
        None, description="Custom minimum discount percentage."
    ),
    stock_filter: str | None = Query(
        None, description="Stock preset (quantity or custom)."
    ),
    stock_min: str | None = Query(None, description="Custom minimum stock value."),
    stock_max: str | None = Query(None, description="Custom maximum stock value."),
    session: Session = Depends(get_session),
):
    """Render a page listing items that became clearance deals today."""

    normalized_state = _normalize_state(state)
    normalized_category = category or None
    filters = _normalize_filters(
        time_window=time_window,
        discount_filter=discount_filter,
        discount_min=discount_min,
        stock_filter=stock_filter,
        stock_min=stock_min,
        stock_max=stock_max,
    )

    return _render_dashboard(
        request,
        scope="new",
        state=normalized_state,
        category=normalized_category,
        filters=filters,
        session=session,
    )


@app.get("/export.xlsx")
def export_excel(
    scope: Scope = Query("all", description="Dataset to export (all or new)."),
    state: str | None = Query(None, description="State filter (WA or OR)."),
    category: str | None = Query(None, description="Optional category filter."),
    time_window: str = Query("all", description="Time filter window key."),
    discount_filter: str | None = Query(
        None, description="Discount preset (percentage or custom)."
    ),
    discount_min: str | None = Query(None, description="Custom minimum discount percentage."),
    stock_filter: str | None = Query(None, description="Stock preset (quantity or custom)."),
    stock_min: str | None = Query(None, description="Custom minimum stock value."),
    stock_max: str | None = Query(None, description="Custom maximum stock value."),
    session: Session = Depends(get_session),
) -> StreamingResponse:
    """Return an Excel workbook for the current filter selection."""

    normalized_state = _normalize_state(state)
    normalized_category = category or None
    filters = _normalize_filters(
        time_window=time_window,
        discount_filter=discount_filter,
        discount_min=discount_min,
        stock_filter=stock_filter,
        stock_min=stock_min,
        stock_max=stock_max,
    )
    raw_items = _select_items(
        session, scope=scope, state=normalized_state, category=normalized_category
    )
    prepared_items = _prepare_listings(raw_items)
    items = _apply_filters(prepared_items, filters=filters)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clearance"
    headers = [
        "Added (UTC)",
        "Last Updated (UTC)",
        "Price Change (UTC)",
        "State",
        "ZIP",
        "Store",
        "SKU",
        "Title",
        "Category",
        "Price",
        "Was",
        "% Off",
        "Availability",
        "Prev Price",
        "Prev Was",
        "Prev Updated (UTC)",
        "URL",
    ]
    sheet.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="004990", end_color="004990", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    sheet.freeze_panes = "A2"

    for item in items:
        pct_off = item.get("pct_off")
        pct_value = pct_off * 100 if pct_off is not None else None
        first_seen = item.get("first_seen") or item.get("price_started_at")
        last_seen = item.get("updated_at") or item.get("price_started_at")
        price_change = item.get("price_started_at")
        prev_updated = item.get("prev_updated_at")
        if isinstance(first_seen, datetime):
            first_seen = first_seen.isoformat()
        if isinstance(last_seen, datetime):
            last_seen = last_seen.isoformat()
        if isinstance(price_change, datetime):
            price_change = price_change.isoformat()
        if isinstance(prev_updated, datetime):
            prev_updated = prev_updated.isoformat()
        sheet.append(
            [
                first_seen,
                last_seen,
                price_change,
                _state_from_zip(item.get("store_zip")),
                item.get("store_zip"),
                item.get("store_name"),
                item.get("sku"),
                item.get("title"),
                item.get("category"),
                item.get("price"),
                item.get("price_was"),
                pct_value,
                item.get("availability"),
                item.get("prev_price"),
                item.get("prev_price_was"),
                prev_updated,
                item.get("product_url"),
            ]
        )

    for column_cells in sheet.iter_cols(
        min_row=1, max_row=sheet.max_row, max_col=sheet.max_column
    ):
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        if max_length <= 0:
            continue
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    state_label = normalized_state or "ALL"
    category_label = (normalized_category or "all").replace(" ", "-")
    filename = f"lowes-{scope}-{state_label.lower()}-{category_label.lower()}.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename={filename}",
    }
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/api/clearance")
def api_clearance(
    scope: Scope = Query("all", description="Dataset to export (all or new)."),
    state: str | None = Query(None, description="State filter (WA or OR)."),
    category: str | None = Query(None, description="Optional category filter."),
    time_window: str = Query("all", description="Time filter window key."),
    discount_filter: str | None = Query(
        None, description="Discount preset (percentage or custom)."
    ),
    discount_min: str | None = Query(None, description="Custom minimum discount percentage."),
    stock_filter: str | None = Query(None, description="Stock preset (quantity or custom)."),
    stock_min: str | None = Query(None, description="Custom minimum stock value."),
    stock_max: str | None = Query(None, description="Custom maximum stock value."),
    session: Session = Depends(get_session),
) -> JSONResponse:
    """Return clearance items as JSON data."""

    normalized_state = _normalize_state(state)
    normalized_category = category or None
    filters = _normalize_filters(
        time_window=time_window,
        discount_filter=discount_filter,
        discount_min=discount_min,
        stock_filter=stock_filter,
        stock_min=stock_min,
        stock_max=stock_max,
    )
    raw_items = _select_items(
        session,
        scope=scope,
        state=normalized_state,
        category=normalized_category,
    )
    prepared_items = _prepare_listings(raw_items)
    items = _apply_filters(prepared_items, filters=filters)
    grouped = _group_listings(items)
    payload = [_serialize_observation(item) for item in items]
    grouped_payload = [_serialize_group(group) for group in grouped]
    filters_payload = {
        "time_window": filters.get("time_window"),
        "time_cutoff": filters.get("time_cutoff").isoformat()
        if filters.get("time_cutoff")
        else None,
        "discount_choice": filters.get("discount_choice"),
        "discount_min_pct": filters.get("discount_min_pct"),
        "stock_choice": filters.get("stock_choice"),
        "stock_min": filters.get("stock_min"),
        "stock_max": filters.get("stock_max"),
    }
    return JSONResponse(
        content={
            "items": payload,
            "groups": grouped_payload,
            "count": len(payload),
            "scope": scope,
            "state": normalized_state or "ALL",
            "category": normalized_category,
            "filters": filters_payload,
        }
    )


@app.get("/api/stats")
def api_stats(session: Session = Depends(get_session)) -> JSONResponse:
    """Return aggregate stats for dashboard clients."""

    payload = _cache_stats(session)
    return JSONResponse(content=payload)


if __name__ == "__main__":
    import uvicorn

    port = int(os.getenv("PORT", "8000"))
    uvicorn.run("app.dashboard:app", host="0.0.0.0", port=port, reload=False)

